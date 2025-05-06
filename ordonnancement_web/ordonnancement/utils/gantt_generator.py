import pandas as pd
from datetime import datetime
from django.db.models import Prefetch
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os
from ..models import Operator, Machine, TaskAssignment

class GanttEngine:
    
    @staticmethod
    def generer_donnees_operateurs():
        """
        Génère les données pour le diagramme de Gantt des opérateurs
        Format de retour :
        [
            {
                'operator_id': str,
                'tasks': [
                    {
                        'task_id': str,
                        'product_id': str,
                        'start': datetime,
                        'end': datetime,
                        'duration': float (minutes),
                        'machine': str
                    },
                    ...
                ]
            },
            ...
        ]
        """
        operators = Operator.objects.prefetch_related(
            Prefetch(
                'taskassignment_set', 
                queryset=TaskAssignment.objects.select_related('task', 'machine')
            )  # <-- Parenthèse fermante manquante ici
        )
        result = []
        
        for operator in operators:
            operator_data = {
                'operator_id': operator.id,
                'tasks': []
            }
            
            for assignment in operator.taskassignment_set.all():
                duration = (assignment.end - assignment.start).total_seconds() / 60
                
                operator_data['tasks'].append({
                    'task_id': assignment.task.id,
                    'product_id': assignment.task.product.id,
                    'start': assignment.start,
                    'end': assignment.end,
                    'duration': round(duration, 2),
                    'machine': assignment.machine.id if assignment.machine else None
                })
            
            result.append(operator_data)
        
        return result

    @staticmethod
    def generer_donnees_machines():
        """
        Génère les données pour le diagramme de Gantt des machines
        Format de retour :
        [
            {
                'machine_id': str,
                'utilization': float (%),
                'tasks': [
                    {
                        'task_id': str,
                        'operator_id': str,
                        'start': datetime,
                        'end': datetime,
                        'product_id': str
                    },
                    ...
                ]
            },
            ...
        ]
        """
        machines = Machine.objects.prefetch_related(
            Prefetch('taskassignment_set',
                    queryset=TaskAssignment.objects.select_related('operator', 'task')))
        
        result = []
        
        for machine in machines:
            assignments = machine.taskassignment_set.all()
            total_time = sum(
                (a.end - a.start).total_seconds() 
                for a in assignments
            ) / 60  # en minutes
            
            machine_data = {
                'machine_id': machine.id,
                'utilization': round((total_time / (8 * 60)) * 100, 2),  # 8h jour
                'tasks': []
            }
            
            for assignment in assignments:
                machine_data['tasks'].append({
                    'task_id': assignment.task.id,
                    'operator_id': assignment.operator.id,
                    'start': assignment.start,
                    'end': assignment.end,
                    'product_id': assignment.task.product.id
                })
            
            result.append(machine_data)
        
        return result

    @staticmethod
    def exporter_vers_excel(gantt_data, filename="gantt_data.xlsx"):
        """
        Exporte les données Gantt vers un fichier Excel avec mise en forme
        Args:
            gantt_data: Données générées par generer_donnees_operateurs/machines
            filename: Chemin du fichier de sortie
        Returns:
            str: Chemin complet du fichier généré
        """
        # Création d'un DataFrame pandas
        df_data = []
        
        for operator in gantt_data:
            for task in operator['tasks']:
                df_data.append({
                    'Opérateur': operator['operator_id'],
                    'Tâche': task['task_id'],
                    'Produit': task['product_id'],
                    'Machine': task['machine'],
                    'Début': task['start'],
                    'Fin': task['end'],
                    'Durée (min)': task['duration']
                })
        
        df = pd.DataFrame(df_data)
        
        # Création du fichier Excel avec openpyxl pour la mise en forme
        wb = Workbook()
        ws = wb.active
        ws.title = "Gantt Opérateurs"
        
        # En-têtes
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = Font(bold=True)
            ws.column_dimensions[col_letter].width = 18
        
        # Données
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                col_letter = get_column_letter(col_num)
                ws[f"{col_letter}{row_num}"] = cell_value
                
                # Formatage des dates
                if headers[col_num-1] in ['Début', 'Fin']:
                    ws[f"{col_letter}{row_num}"].number_format = 'DD/MM/YYYY HH:MM'
        
        # Sauvegarde
        abs_path = os.path.abspath(filename)
        wb.save(filename)
        return abs_path